# coding: utf-8
# -*- coding: utf-8 -*-

# This file is part of the FlOpEDT/FlOpScheduler project.
# Copyright (c) 2017
# Authors: Iulian Ober, Paul Renaud-Goud, Pablo Seban, et al.
# 
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU Affero General Public License as
# published by the Free Software Foundation, either version 3 of the
# License, or (at your option) any later version.
# 
# This program is distributed in the hope that it will be useful, but
# WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the GNU
# Affero General Public License for more details.
# 
# You should have received a copy of the GNU Affero General Public
# License along with this program. If not, see
# <http://www.gnu.org/licenses/>.
# 
# You can be released from the requirements of the license by purchasing
# a commercial license. Buying such a license is mandatory as soon as
# you develop activities involving the FlOpEDT/FlOpScheduler software
# without disclosing the source code of your own applications.

import os
import json
import logging

from django.http import HttpResponse
from django.shortcuts import render
from django.db import transaction
from django.conf import settings

from core.decorators import dept_admin_required

from base.models import Department, Period, Week

from configuration.make_planif_file import make_planif_file
from configuration.make_filled_database_file import make_filled_database_file
from configuration.extract_planif_file import extract_planif
from configuration.deploy_database import extract_database_file
from configuration.file_manipulation import upload_file, check_ext_file
from configuration.forms import ImportPlanif, ImportConfig
from base.weeks import current_year

logger = logging.getLogger(__name__)

@dept_admin_required
def configuration(req, **kwargs):
    """
    Main view of Configuration
    :param req:
    :return:
    """
    arg_req = {}

    arg_req['form_config'] = ImportConfig()
    arg_req['form_planif'] = ImportPlanif()

    arg_req['departements'] = [{'name': depart.name, 'abbrev': depart.abbrev}
                               for depart in Department.objects.all() if not depart.abbrev == 'default']
    arg_req['periods'] = [{'name': period.name, 'department': period.department.abbrev}
                          for period in Period.objects.all()]
    arg_req['current_year'] = current_year
    return render(req, 'configuration/configuration.html', arg_req)


@dept_admin_required
def import_config_file(req, **kwargs):
    """
    View for the first step of the configuration. It imports the file
    to build the database, clear the database, extract the file to
    build the database and make the planif file for the second step
    of the configuration.
    Ajax request.

    :param req:
    :return:
    """
    if req.method == 'POST':
        form = ImportConfig(req.POST, req.FILES)
        logger.debug(req)
        logger.debug(req.FILES)
        if form.is_valid():
            logger.debug(req.FILES['fichier'])
            logger.debug(req.FILES['fichier'].name)
            if check_ext_file(req.FILES['fichier'], ['.xlsx', '.xls']):
                path = upload_file(req.FILES['fichier'], "uploaded_database_file.xlsx")
                # If one of method fail the transaction will be not commit.
                try:
                    with transaction.atomic():
                        dept_abbrev = req.POST['abbrev']
                        try:
                            dept_name = req.POST['name']
                        except:
                            dept_name = None
                        logger.debug(dept_name)
                        try:
                            dept = Department.objects.get(abbrev=dept_abbrev)
                            if not dept_name == dept.name and dept_name is not None:
                                response = {'status': 'error',
                                            'data': "Il existe déjà un département utilisant cette abbréviation."}
                                return HttpResponse(json.dumps(response), content_type='application/json')
                            dept_name = dept.name
                            dept.delete()
                            logger.debug("flush OK")
                        except Exception as e:
                            logger.warning(f'Exception with dept')
                            logger.warning(e)

                        extract_database_file(department_name=dept_name,
                                              department_abbrev=dept_abbrev, bookname=path)
                        logger.debug("extract OK")

                        os.rename(path, os.path.join(settings.MEDIA_ROOT,
                                                     'configuration',
                                                     f'database_file_{dept_abbrev}.xlsx'))
                        logger.warning("rename OK")
                        response = {'status': 'ok',
                                    'data': 'OK',
                                    'dept_abbrev': dept_abbrev,
                                    'dept_fullname': dept_name
                        }
                except Exception as e:
                    os.remove(path)
                    logger.debug(e)
                    response = {'status': 'error', 'data': str(e)}
                    return HttpResponse(json.dumps(response), content_type='application/json')
                dept = Department.objects.get(abbrev=dept_abbrev)
                source = os.path.join(settings.MEDIA_ROOT,
                                      'configuration',
                                      'empty_planif_file.xlsx')
                target_repo = os.path.join(settings.MEDIA_ROOT,
                                           'configuration')
                logger.info("start planif")
                make_planif_file(dept, empty_bookname=source, target_repo=target_repo)
                logger.info("make planif OK")
            else:
                response = {'status': 'error', 'data': 'Invalid format'}
        else:
            response = {'status': 'error', 'data': 'Form not valid'}
    return HttpResponse(json.dumps(response), content_type='application/json')


@dept_admin_required
def get_config_file(req, **kwargs):
    """
    Resend the empty configuration's file.

    :param req:
    :return:
    """
    f = open(f"{settings.MEDIA_ROOT}/configuration/empty_database_file.xlsx", "rb")
    response = HttpResponse(f, content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="database_file.xls"'
    f.close()
    return response


@dept_admin_required
def get_planif_file(req, with_courses=False, **kwargs):
    """
    Send an empty planification's file.
    Rely on the configuration step if it was taken.
    :param req:
    :return:
    """
    logger.debug(req.GET['departement'])
    filename = os.path.join(settings.MEDIA_ROOT,
                             'configuration',
                             f"planif_file_{req.GET['departement']}")
    if with_courses:
        filename += '_with_courses'
    filename += ".xlsx"

    if not os.path.exists(filename):
        filename = os.path.join(settings.MEDIA_ROOT,
                                'configuration',
                                f"empty_planif_file.xlsx")
    f = open(filename, "rb")
    response = HttpResponse(f, content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="planif_file.xlsx"'
    f.close()
    return response

@dept_admin_required
def get_filled_database_file(req, **kwargs):
    """
    Send an filled database file.
    Rely on the configuration step if it was taken.
    :param req:
    :return:
    """
    logger.debug(req.GET['departement'])
    basic_filename = f"database_file_{req.GET['departement']}"
    filename = os.path.join(settings.MEDIA_ROOT,
                             'configuration',
                             basic_filename)
    filename += ".xlsx"

    if not os.path.exists(filename):
        filename = os.path.join(settings.MEDIA_ROOT,
                                'configuration',
                                f"empty_database_file.xlsx")
    f = open(filename, "rb")
    response = HttpResponse(f, content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="planif_file.xlsx"'
    f.close()
    return response


@dept_admin_required
def mk_and_dl_planif(req, with_courses, **kwargs):
    logger.debug(req.GET['departement'])
    dept_abbrev = req.GET['departement']
    dept = Department.objects.get(abbrev=dept_abbrev)
    source = os.path.join(settings.MEDIA_ROOT,
                          'configuration',
                          'empty_planif_file.xlsx')
    target_repo = os.path.join(settings.MEDIA_ROOT,
                               'configuration')
    logger.info("start planif")
    make_planif_file(dept, empty_bookname=source, target_repo=target_repo, with_courses=with_courses)
    return get_planif_file(req, with_courses, **kwargs)


@dept_admin_required
def mk_and_dl_database_file(req, **kwargs):
    logger.debug(req.GET['departement'])
    dept_abbrev = req.GET['departement']
    dept = Department.objects.get(abbrev=dept_abbrev)
    logger.info("start filled database file")
    make_filled_database_file(dept)
    return get_filled_database_file(req, **kwargs)


@dept_admin_required
def import_planif_file(req, **kwargs):
    """
    Import a planification's file filled. Before data processing, it must to
    check if the first step of te configuration is done. Extract the data of the xlsx file.

    :param req:
    :return:
    """
    form = ImportPlanif(req.POST, req.FILES)
    if form.is_valid():
        if check_ext_file(req.FILES['fichier'], ['.xlsx', '.xls']):
            logger.info(req.FILES['fichier'])
            path = upload_file(req.FILES['fichier'], "configuration/planif_file_.xlsx")
            # If one of methods fail, the transaction will be not commit.
            try:
                with transaction.atomic():
                    try:
                        dept = Department.objects.get(abbrev=req.POST['departement'])
                    except Exception as e:
                        response = {'status': 'error', 'data': str(e)}
                        return HttpResponse(json.dumps(response), content_type='application/json')
                    stabilize_courses = "stabilize" in req.POST
                    assign_colors = "assign_colors" in req.POST
                    print("AAAAA", assign_colors)
                    choose_weeks = "choose_weeks" in req.POST
                    choose_periods = "choose_periods" in req.POST
                    if choose_weeks:
                        week_nb = req.POST["week_nb"]
                        year = req.POST["year"]
                        if not week_nb and not year:
                            from_week = None
                        else:
                            week_nb = int(week_nb)
                            year = int(year)
                            from_week = Week.objects.get(nb=week_nb, year=year)
                        week_nb_end = req.POST["week_nb_end"]
                        year_end = req.POST["year_end"]
                        if not week_nb_end and not year_end:
                            until_week = None
                        else:
                            week_nb_end = int(week_nb_end)
                            year_end = int(year_end)
                            until_week = Week.objects.get(nb=week_nb_end, year=year_end)
                    else:
                        from_week = None
                        until_week = None

                    if choose_periods:
                        periods = Period.objects.filter(department=dept, name__in=req.POST.getlist('periods'))
                        print(periods)
                    else:
                        periods = None

                    extract_planif(dept, bookname=path, from_week=from_week, until_week=until_week, periods=periods,
                                   stabilize_courses=stabilize_courses, assign_colors=assign_colors)
                    logger.info("Extract file OK")
                    rep = "OK !"

                    os.rename(path, f"{settings.MEDIA_ROOT}/configuration/planif_file.xlsx")
                    logger.info("Rename OK")

                    response = {'status': 'ok', 'data': rep}
            except Exception as e:
                os.remove(path)
                logger.info(e)
                response = {'status': 'error', 'data': str(e)}
        else:
            response = {'status': 'error', 'data': 'Invalid format'}
    else:
        response = {'status': 'error', 'data': "Form can't be valid"}
    return HttpResponse(json.dumps(response), content_type='application/json')


@dept_admin_required
def export_to_excel(req, **kwargs):


    arg_req = {}

    arg_req['form_config'] = "test"
    arg_req['form_planif'] = "test"

    arg_req['departements'] = [{'name': depart.name, 'abbrev': depart.abbrev}
                               for depart in Department.objects.all() if not depart.abbrev == 'default']
    arg_req['current_year'] = current_year

    return render(req, 'configuration/export_to_excel.html', arg_req)

def createExportExcel(req, **kwargs):
    
    import openpyxl
    from openpyxl.styles import PatternFill
    from openpyxl.drawing.image import Image

    #Récupération des valeurs choisies coté client
    valueGroupe = req.GET.get('groupe')
    valueAnnee = req.GET.get('annee')
    valueSemestre = req.GET.get('semestre')

    #------------------INFOS MODULABLES-------------------#
    numSemaineDebut, numSemaineFin = getBeginAndEndWeeks(valueSemestre)
    numDerniereSemaineAnnee = 52 

    departement = valueGroupe
    anneeEtu = valueAnnee
    semestre = valueSemestre
    anneeScolaire = 2023 #2023 pour 2022-2023

    dept = "HES"
    tree = getTreeGroup(dept)
    promo = valueAnnee

    for j in range(len(tree)):
        if tree[j]["name"] == promo :
            subTree = tree[j]
            
    for j in range(len(subTree["children"])):
        if subTree["children"][j]["name"] == valueGroupe :
            subSubTree = subTree["children"][j]

    listeDeGroupe = []
    for j in range(len(subSubTree["children"])):
        listeDeGroupe.append(subSubTree["children"][j]["name"])

    #------------ECRITURE DU FICHIER-------------#

    # Créez un nouveau classeur
    wb = openpyxl.Workbook()

    # Créez une nouvelle feuille de calcul
    sheet = wb.create_sheet("Semestre "+str(semestre)+" "+str(anneeScolaire)+"-"+str(anneeScolaire+1))

    #index de la premiere colonne du Excel (3 = C)
    start_column_index = 3

    #nombre de semaines
    boolNewYear = (numSemaineDebut - numSemaineFin) > 0
    if boolNewYear :
        indiceBeginning = numSemaineDebut
        indiceEnding = numDerniereSemaineAnnee +  numSemaineFin
    else :
        indiceBeginning = numSemaineDebut
        indiceEnding = numSemaineFin

    for i in range(indiceBeginning, indiceEnding + 1):
        i = i % numDerniereSemaineAnnee
        if i == 0 :
            i = numDerniereSemaineAnnee

        start_column = openpyxl.utils.get_column_letter(start_column_index)
        end_column = openpyxl.utils.get_column_letter(start_column_index + len(listeDeGroupe)-1)
        
        cells_to_merge = f'{start_column}{3}:{end_column}{3}'
        sheet.merge_cells(cells_to_merge)

        #--------------CELLULE DE LA SEMAINE S-xx-----------------#
        #pour avoir la derniere bordure à droite
        borders = openpyxl.styles.Border(
            top=None,
            left=None,
            right=openpyxl.styles.Side(border_style='thick'),
            bottom=None)
        sheet[f'{end_column}{3}'].border = borders

        #Cellule à modifier
        cellActuelle = sheet[f'{start_column}{3}']
        #taille et police d'ecriture
        font_size = 22
        font_name = 'Arial'
        cellActuelle.font = openpyxl.styles.Font(size=font_size, name=font_name)
        #centrer le texte
        cellActuelle.alignment = openpyxl.styles.Alignment(horizontal='center', vertical = 'center')
        #bordure de la case
        borders = openpyxl.styles.Border(
            top=openpyxl.styles.Side(border_style='thick'),
            left=openpyxl.styles.Side(border_style='thick'),
            right=None,
            bottom=None)
        cellActuelle.border = borders
        #Texte dans la cellule
        sheet[f'{start_column}{3}'].value = 'S' + str(i)
    
        row = sheet.row_dimensions[3]
        heightLine3and4 = len(cellActuelle.value.split('\n')) * 22
        row.height = heightLine3and4

        #-----cellule contenant les dates debut et fin de la semaine------#
        dateDebut, dateFin = findBeginAndEndOfWeek(anneeScolaire, i)
        row = sheet.row_dimensions[4]
        row.height = heightLine3and4

        #cellule date debut
        cellActuelle = sheet[f'{start_column}{4}']
        borders = openpyxl.styles.Border(
            top=None,
            left=openpyxl.styles.Side(border_style='thick'),
            right=None,
            bottom=openpyxl.styles.Side(border_style='thick'))
        cellActuelle.border = borders
        font_size = 11
        font_name = 'Arial'
        cellActuelle.font = openpyxl.styles.Font(size=font_size, name=font_name)
        cellActuelle.alignment = openpyxl.styles.Alignment(horizontal='left', vertical = 'center')
        sheet[f'{start_column}{4}'] = dateDebut
        column = sheet.column_dimensions[f'{start_column}']
        column.width = len(cellActuelle.value) * 1.3

        #cellule date fin
        cellActuelle = sheet[f'{end_column}{4}']
        borders = openpyxl.styles.Border(
            top=None,
            left=None,
            right=openpyxl.styles.Side(border_style='thick'),
            bottom=openpyxl.styles.Side(border_style='thick'))
        cellActuelle.border = borders
        font_size = 11
        font_name = 'Arial'
        cellActuelle.font = openpyxl.styles.Font(size=font_size, name=font_name)
        cellActuelle.alignment = openpyxl.styles.Alignment(horizontal='right', vertical = 'center')
        sheet[f'{end_column}{4}'] = dateFin
        column = sheet.column_dimensions[f'{end_column}']
        column.width = len(cellActuelle.value) * 1.3
        
        #groupes pour chaque semaine
        for j in range(len(listeDeGroupe)):
            index = openpyxl.utils.get_column_letter(start_column_index + j)
            cellActuelle = sheet[f'{index}{5}']
            sheet[f'{index}{5}'] = listeDeGroupe[j]
            borders = openpyxl.styles.Border(
                top=openpyxl.styles.Side(border_style='thick'),
                left=openpyxl.styles.Side(border_style='thick'),
                right=openpyxl.styles.Side(border_style='thick'),
                bottom=openpyxl.styles.Side(border_style='thick'))
            cellActuelle.border = borders
            font_size = 14
            font_name = 'Arial'
            cellActuelle.font = openpyxl.styles.Font(size=font_size, name=font_name, bold=True)
            cellActuelle.alignment = openpyxl.styles.Alignment(horizontal='center', vertical = 'center')
            new_line_height = font_size * 3.5
            sheet.row_dimensions[sheet[f'{index}{5}'].row].height = new_line_height

        #DEBUT DE LA REQUEST POUR LES COURS DE LA SEMAINE-------------------------------------#
        week = i
        year = anneeScolaire + (i < numSemaineDebut and boolNewYear)
        
        #recup des sous groupes du groupe demandé
        scheduledTotal = []

        for j in range(len(listeDeGroupe)):
            scheduled = getScheduledByDeptWeekYear(dept, str(week), str(year), promo, listeDeGroupe[j])
            print(scheduled)
            scheduledTotal += scheduled

        #TRI DES COURS DANS UN TABLEAU DE LA MEME FORME QU'UNE COLONNE-SEMAINE DANS LE EXCEL--#
        orderTabCourse = orderScheduledWeek(scheduledTotal, listeDeGroupe)
        

        #Completion du excel avec les cours
        for j in range (10): #pour chaque demi-journée
            for k in range (3): #pour chaque créneau d'une demi-journée
                for l in range (len(listeDeGroupe)): #pour chaque colonne de groupe
                    index = openpyxl.utils.get_column_letter(start_column_index + l)
                    cellActuelle = sheet[f'{index}{7 + j * 4 + k}']
                    sheet[f'{index}{7 + j * 4 + k}'] = orderTabCourse[l]["cours"][j*3 + k]["name"]
                    fill = PatternFill(start_color=orderTabCourse[l]["cours"][j*3 + k]["color"], end_color=orderTabCourse[l]["cours"][j*3 + k]["color"], fill_type='solid')
                    cellActuelle.fill = fill
                    if (l == 0):
                        borders = openpyxl.styles.Border(
                            top=openpyxl.styles.Side(border_style='thin'),
                            left=openpyxl.styles.Side(border_style='thick'),
                            right=openpyxl.styles.Side(border_style='thin'),
                            bottom=openpyxl.styles.Side(border_style='thin'))
                    elif (l == len(listeDeGroupe)-1):
                        borders = openpyxl.styles.Border(
                            top=openpyxl.styles.Side(border_style='thin'),
                            left=openpyxl.styles.Side(border_style='thin'),
                            right=openpyxl.styles.Side(border_style='thick'),
                            bottom=openpyxl.styles.Side(border_style='thin'))
                    else:
                        borders = openpyxl.styles.Border(
                            top=openpyxl.styles.Side(border_style='thin'),
                            left=openpyxl.styles.Side(border_style='thin'),
                            right=openpyxl.styles.Side(border_style='thin'),
                            bottom=openpyxl.styles.Side(border_style='thin'))
                    cellActuelle.border = borders
                    cellActuelle.alignment = openpyxl.styles.Alignment(wrapText=True, horizontal='center', vertical = 'center')
                    #mise a jour de la hauteur de la cellule  --> /!\ IL FAUT GARDER LA PLUS GRANDE HAUTEUR AU FUR ET A MESURE DE LA CONSTRUCTION
                    font = cellActuelle.font
                    font_size = font.size
                    new_line_height = font_size * 3.5 # 1 point = 0.035 cm
                    actualRowHeight = sheet.row_dimensions[cellActuelle.row].height
                    
                    if (actualRowHeight == None or actualRowHeight < new_line_height) :
                        sheet.row_dimensions[cellActuelle.row].height = new_line_height

        #colonne suivante (semaine suivante)
        start_column_index += len(listeDeGroupe)

    #--------------------LIGNES HORIZONTALES--------------------#

    #entre chaque jour (en noir)
    for i in range (5):
        end_column = openpyxl.utils.get_column_letter(start_column_index - 1)
        cells_to_merge = f'{"A"}{14 + i*8}:{end_column}{14 + i*8}'
        sheet.merge_cells(cells_to_merge)
        cell = sheet[f'{"A"}{14 + i*8}']
        cell.fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=openpyxl.styles.colors.Color(rgb='000000'))
        
    #pause du midi (alterne bleu foncé et clair)
    for i in range (5):
        end_column = openpyxl.utils.get_column_letter(start_column_index - 1)
        cells_to_merge = f'{"B"}{10 + i*8}:{end_column}{10 + i*8}'
        sheet.merge_cells(cells_to_merge)
        cell = sheet[f'{"B"}{10 + i*8}']
        if (i % 2 == 0):
            cell.fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=openpyxl.styles.colors.Color(rgb='0066CC'))
        else:
            cell.fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=openpyxl.styles.colors.Color(rgb='33CCCC'))
    #-----------------------------------------------------------#

    #---------------ECRITURE DES JOURS EN VERTICAL---------------#

    jours = ["LUNDI", "MARDI", "MERCREDI", "JEUDI", "VENDREDI"]

    for i in range (5):
        cells_to_merge = f'{"A"}{7 + i*8}:{"A"}{13 + i*8}'
        sheet.merge_cells(cells_to_merge)
        cellActuelle = sheet[f'{"A"}{7 + i*8}']
        cellActuelle.alignment = openpyxl.styles.Alignment(textRotation=90, horizontal="center", vertical="center")
        font_size = 22
        font_name = 'Arial'
        cellActuelle.font = openpyxl.styles.Font(size=font_size, name=font_name, bold=True)
        sheet[f'{"A"}{7 + i*8}'] = jours[i]
        borders = openpyxl.styles.Border(
            top=openpyxl.styles.Side(border_style='thin'),
            left=openpyxl.styles.Side(border_style='thin'),
            right=openpyxl.styles.Side(border_style='thin'),
            bottom=openpyxl.styles.Side(border_style='thin'))
        cellActuelle.border = borders
    #------------------------------------------------------------#

    #---------------ECRITURE DES HEURES DE CRENEAUX COURS---------------#
    heuresCreneaux = ["8h\n9h30", "9h45\n11h15","11h30\n13h","14h\n15h30","15h45\n17h15","17h30\n19h"]
    for i in range ((len(heuresCreneaux))):
        if (i < 3):
            for j in range (5):
                cellActuelle = sheet[f'{"B"}{7+i + j*8}']
                sheet[f'{"B"}{7+i + j*8}'] = heuresCreneaux[i]
                borders = openpyxl.styles.Border(
                    top=openpyxl.styles.Side(border_style='thin'),
                    left=openpyxl.styles.Side(border_style='thin'),
                    right=openpyxl.styles.Side(border_style='thin'),
                    bottom=openpyxl.styles.Side(border_style='thin'))
                cellActuelle.border = borders
                cellActuelle.alignment = openpyxl.styles.Alignment(wrapText=True, horizontal='center', vertical = 'center')

        #passer la ligne du midi de chaque jour
        else:
            for j in range (5):
                cellActuelle = sheet[f'{"B"}{8+i + j*8}']
                sheet[f'{"B"}{8+i + j*8}'] = heuresCreneaux[i]
                borders = openpyxl.styles.Border(
                    top=openpyxl.styles.Side(border_style='thin'),
                    left=openpyxl.styles.Side(border_style='thin'),
                    right=openpyxl.styles.Side(border_style='thin'),
                    bottom=openpyxl.styles.Side(border_style='thin'))
                cellActuelle.border = borders
                cellActuelle.alignment = openpyxl.styles.Alignment(wrapText=True, horizontal='center', vertical = 'center')

    #---------------------AJOUT DU LOGO POLYTECH----------------------#
    #problème ici : le fichier n'arrive pas a trouver l'image dans les fichiers flop
    #image = Image('./../img/logoPolytechNantes.png')
    #image.left = 10
    #image.top = 10
    #sheet.add_image(image) #décale de 10px vers bas droite, en partant de haut Gauche A1

    #---------------------AJOUT DU TITRE DU EXCEL ----------------------#
    font_size = 26
    font_name = 'Arial'

    indexNumber = int((start_column_index - 2) / 2 - len(listeDeGroupe) * 2) + 1 #approximativement le milieu
    indexTitle = openpyxl.utils.get_column_letter(indexNumber) 

    title = "Emploi du temps de : " + str(departement) + " / Année " + str(anneeEtu)
    sheet[f'{indexTitle}{1}'].font = openpyxl.styles.Font(size=font_size, name=font_name, bold=True)
    sheet[f'{indexTitle}{1}'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical = 'center')
    sheet[f'{indexTitle}{1}'] = title

    yearAndSemester = str(anneeScolaire) + "-" + str(anneeScolaire+1) + "   Semestre " + str(semestre)
    sheet[f'{indexTitle}{2}'].font = openpyxl.styles.Font(size=font_size, name=font_name, bold=True)
    sheet[f'{indexTitle}{2}'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical = 'center')
    sheet[f'{indexTitle}{2}'] = yearAndSemester

    new_line_height = font_size * 3.5
    sheet.row_dimensions[sheet[f'{indexTitle}{1}'].row].height = new_line_height
    sheet.row_dimensions[sheet[f'{indexTitle}{2}'].row].height = new_line_height

    #----------------------FIN DE LA CREATION DU TEMPLATE ---------------------#
    # Définissez le zoom de la feuille sur 60 %
    sheet.sheet_view.zoomScale = 60
    # Feuille par défaut
    sheet = wb['Sheet']
    # Supprimez la feuille de calcul
    wb.remove(sheet)

    # Return the response value 
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="mon_fichier.xlsx"'

    wb.save(response)
    return response

def findBeginAndEndOfWeek(year, week_number):

        from datetime import datetime, timedelta
        mois = ['Janv.', 'Févr.', 'Mars', 'Avr.', 'Mai', 'Juin', 'Juil.', 'Août', 'Sept.', 'Oct.', 'Nov.', 'Déc.']

        # Calculez le premier jour de la semaine sélectionnée
        first_day_of_week = datetime(year, 1, 1) + timedelta((week_number) * 7 - (datetime(year, 1, 1).weekday() + 1) % 7)

        # Calculez la date de début et de fin de la semaine en enlevant les jours du week-end
        start_date = first_day_of_week + timedelta(days=1)
        end_date = first_day_of_week + timedelta(days=5)

        # Affichez la date de début et de fin de la semaine
        dateDebut = str(start_date.day) + " " + mois[start_date.month - 1]
        dateFin = str(end_date.day) + " " + mois[end_date.month - 1]
        
        return dateDebut, dateFin

def getTreeGroup(dept):
    #permet d'obtenir l'arborescence des groupes/sous-groupes du département en paramètre
    import requests
    request = 'http://localhost:8000/fr/api/groups/structural/tree/?dept=' + dept

    r = requests.get(request)
    content = r.content
    data = json.loads(content.decode('utf-8'))
    if (r.status_code == 200):
        return data
    else:
        return "Error, request status : ", r.status_code

def getScheduledByDeptWeekYear(dept, week, year, promo, group):
    #permet d'obtenir l'ensemble des cours sur une semaine donnée
    import requests

    request='http://localhost:8000/fr/api/fetch/scheduledcourses/?week='+week
    +'&year='+year
    +'&dept='+dept
    +'&train_prog='+promo
    +'&group='+group

    r = requests.get(request)
    content = r.content
    data = json.loads(content.decode('utf-8'))
    if (r.status_code == 200):
        return data
    else:
        return "Error, request status : ", r.status_code

def orderScheduledWeek(scheduled, listOfGroupe):
    #fonction permettant de formater la liste de cours récupéré via l'API en un tableau réduit
    heuresCreneaux = [480, 570, 660, 840, 930, 1020]
    jourAbbrev = {
    "m" : 0,
    "tu" : 1, 
    "w" : 2,
    "th" : 3, 
    "f" : 4,
    }
    result = []
    for i in range (len(listOfGroupe)):
        tabTemp = {
            'groupe' : listOfGroupe[i],
            'cours': []
            }
        for j in range(30):
            tabTemp["cours"].append({"name" : "", "color" : "FFFFFF"})
        result.append(tabTemp)
    
    for i in range (len(scheduled)):
        thisCourse = scheduled[i]
        #recup heure de debut
        fullMinutes = thisCourse["start_time"]
        #hourBegin = fullMinutes // 60
        #minuteBegin = fullMinutes % 60
        indexCreneau = findBeginCoursIndex(heuresCreneaux, fullMinutes)
        abbrevDay = str(thisCourse["day"])
        indexJour = jourAbbrev[abbrevDay]
        nameCours = thisCourse["course"]["module"]["abbrev"]
        colorCours = thisCourse["course"]["module"]["display"]["color_bg"][1:]

        #recup les groupes qui participent a ce cours
        for j in range(len(thisCourse["course"]["groups"])):
            for k in range (len(result)):
                if result[k]["groupe"] == thisCourse["course"]["groups"][j]["name"] :
                    result[k]["cours"][indexJour * 6 + indexCreneau] = {"name" : nameCours, "color" : colorCours}
    return result

def findBeginCoursIndex(tableau, value):
    #permet de connaitre le créneau d'un cours via l'heure du début
    for i in range(len(tableau)-1, -1, -1):
        if tableau[i] <= value:
            return i
    return None

def getBeginAndEndWeeks(semester):
    #marche uniquement pour HES (departement:33) pour l'instant 
    import requests

    request = 'http://localhost:8000/fr/api/base/periods/'

    r = requests.get(request)
    content = r.content
    data = json.loads(content.decode('utf-8'))
    if (r.status_code == 200):
        for i in data :
            if i['department']==33 and i['name'] == "S"+semester:
                return i['starting_week'], i['ending_week']
    else:
        return "Error, request status : ", r.status_code