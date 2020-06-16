#!/usr/bin/env python3
# -*- coding: utf-8 -*-

# This file is part of the FlOpEDT/FlOpScheduler project.
# Copyright (c) 2017
# Authors: Iulian Ober, Paul Renaud-Goud, Pablo Seban, et al.
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU Affero General Public License as
# published by the Free Software Foundation, either version 3 of the
# License, or (at your option) any later version.
#
# This program is distributed in the hope that it will be useful, but
# WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the GNU
# Affero General Public License for more details.
#
# You should have received a copy of the GNU Affero General Public
# License along with this program. If not, see
# <http://www.gnu.org/licenses/>.
#
# You can be released from the requirements of the license by purchasing
# a commercial license. Buying such a license is mandatory as soon as
# you develop activities involving the FlOpEDT/FlOpScheduler software
# without disclosing the source code of your own applications.

# This code takes a XLSX database file description and turns it into
# structured Python data, with the main parse_file function, following
# the data structure described in database_description_checker.py.
#

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

import logging
logger = logging.getLogger(__name__)

people_sheet = 'Intervenants'
rooms_sheet = 'Salles'
groups_sheet = 'Groupes'
modules_sheet = 'Modules'
courses_sheet = 'Cours'
settings_sheet = 'Paramètres'

REASONABLE = 3141 # large enough?

# trivial helper
def cell_name(row, column):
    return '{0:s}{1:d}'.format(get_column_letter(column), row)

#################################################
#                                               #
#   Helper functions to parse various types     #
#   of cells, either individually or by range   #
#                                               #
#################################################

def parse_integer(sheet, row, column):
    try:
        return int(sheet.cell(row=row, column=column).value)
    except:
        return None

def parse_time(sheet, row, column):
    "Helper function to get a time out of a cell"
    "as a number of minutes since midnight"
    "(will return None if anything goes wrong)"
    try:
        val = sheet.cell(row=row, column=column).value
        return 60 * val.hour + val.minute
    except:
        return None

def parse_time_list_in_line(sheet, row, col_start):
    "Parse a line representing a list of times"
    "(stop at the first empty cell)"

    result = []
    col = col_start
    while col < REASONABLE:
        val = parse_time(sheet, row, col)
        if val == None:
            break
        result.append(val)
        col = col + 1

    return result

def parse_string(sheet, row, column):
    "Helper function to get a clean string out of a cell"
    "(will return '' if there's nothing to see)"

    val = sheet.cell(row=row, column=column).value
    if val == None:
        val=''
    val = str(val).strip()

    return val

def parse_string_set_in_line(sheet, row, col_start):
    "Parse a line representing a set of strings"
    "(stop at the first empty cell)"

    result = set()
    col = col_start
    while col < REASONABLE:
        val = parse_string(sheet, row, col)
        if val == '':
            break
        result.add(val)
        col = col + 1

    return result

def parse_string_set_dictionary(sheet, row_start, col_start, row_end = REASONABLE):
    "Parse a block, turning it into a dictionary of string sets"
    "The first column gives the keys, the line at the right of the key is"
    "the associated value"
    result = dict()
    row = row_start
    while row < row_end:
        name = parse_string(sheet, row, col_start)
        if name == '':
            row = row + 1
            continue
        if name in result:
            name = ':INVALID:DUPLICATE:{0:s}'.format(cell_name(row, col_start))
        result[name] = parse_string_set_in_line(sheet, row, col_start + 1)
        row = row + 1

    return result

################################
#                              #
#   Various helper functions   #
#                              #
################################


def find_marker_cell(sheet, marker, row = 1, col = 1):
    "Helper function to find the marker of a data block"
    "(Will return either row, col or None, None)"
    "(With optional parameters to start the search at some position)"

    while row < REASONABLE:
        while col < REASONABLE:
            if parse_string(sheet, row, col) == marker and sheet.cell(row=row, column=col).font.size > 10:
                return row, col
            col = col + 1
        row = row + 1
        col = 1
    return None, None

#################################################
#                                               #
#   Parser functions for the different pages    #
#                                               #
#################################################

def parse_rooms(sheet):
    row_groups, col_groups = find_marker_cell(sheet, 'Groupes')
    row_cats, col_cats = find_marker_cell(sheet, 'Catégories')
    if col_groups != col_cats or row_cats < row_groups:
        logger.warning(f"The marker cells in sheet {rooms_sheet} are misplaced")
        return set(), dict(), dict()

    #
    # parse the groups
    #
    groups = parse_string_set_dictionary(sheet, row_groups + 1, col_groups, row_cats)

    #
    # parse the categories
    #
    categories = parse_string_set_dictionary(sheet, row_cats + 1, col_cats)

    #
    # Build the set of rooms
    #
    # FIXME: what if a room group or category gets named like a room?
    rooms = set()
    for lst in groups.values():
        rooms.update(lst)
    for lst in categories.values():
        rooms.update(lst)
    rooms.difference_update(groups.keys())
    rooms.difference_update(categories.keys())

    return rooms, groups, categories

def parse_people(sheet):
    row, col = find_marker_cell(sheet, "Identifiant")
    if row == None:
        return dict()

    row = row + 1
    result = dict()
    while row < REASONABLE:
        id_ = parse_string(sheet, row, col)
        if id_ == '':
            row = row + 1
            continue
        if id_ in result:
            id_ = ':INVALID:DUPLICATE:{0:s}'.format(cell_name(row, col))
        result[id_] = {'first_name': parse_string(sheet, row, col + 1),
                       'last_name': parse_string(sheet, row, col + 2),
                       'email': parse_string(sheet, row, col + 3),
                       'status': parse_string(sheet, row, col + 4),
                       'employer': parse_string(sheet, row, col + 5)}
        row = row + 1
    return result

def parse_modules(sheet):
    row, col = find_marker_cell(sheet, "Identifiant")
    if row == None:
        logger.warning(f"The marker cell in sheet {modules_sheet} is missing")
        return dict()

    row = row + 1
    result = dict()
    while row < REASONABLE:
        id_ = parse_string(sheet, row, col)
        if id_ == '':
            row = row + 1
            continue
        if id_ in result:
            id_ = ':INVALID:DUPLICATE:{0:s}'.format(cell_name(row, col))
        result[id_] = { 'short': parse_string(sheet, row, col + 1),
                        'PPN': parse_string(sheet, row, col + 2),
                        'name': parse_string(sheet, row, col + 3),
                        'promotion': parse_string(sheet, row, col + 4),
                        'period': parse_string(sheet, row, col + 5),
                        'responsable': parse_string(sheet, row, col + 6)}
        row = row + 1
    return result

def parse_courses(sheet):
    row, col = find_marker_cell(sheet, 'Type')
    if row == None:
        logger.warning(f"The marker cell in sheet {courses_sheet} is missing")
        return dict()

    row = row + 1
    result = dict()
    while row < REASONABLE:
        id_ = parse_string(sheet, row, col)
        if id_ == '':
            row = row + 1
            continue
        if id_ in result:
            id_ = ':INVALID:DUPLICATE:{0:s}'.format(cell_name(row, col))
        try:
            duree = int(parse_string(sheet, row, col + 1))
        except:
            duree = -1
        result[id_] = {'duration': duree,
                       'group_types': set(parse_string_set_in_line(sheet, row, col + 2)),
                       'start_times': set(parse_time_list_in_line(sheet, row + 1, col + 2))}
        row = row + 2 # sic
    return result

def parse_settings(sheet):
    result = dict()
    row, col = find_marker_cell(sheet, 'Jalon')
    if row == None:
        logger.warning(f"The 'Jalon' cell in sheet {settings_sheet} is missing")
        result['day_start_time'] = -1
        result['day_finish_time'] = -1
        result['lunch_break_start_time'] = -1
        result['lunch_break_finish_time'] = -1
    else:
        val = parse_time(sheet, row + 1, col + 1)
        if val == None:
            val = -1
        result['day_start_time'] = val
        val = parse_time(sheet, row + 2, col + 1)
        if val == None:
            val = -1
        result['day_finish_time'] = val
        val = parse_time(sheet, row + 3, col + 1)
        if val == None:
            val = -1
        result['lunch_break_start_time'] = val
        val = parse_time(sheet, row + 4, col + 1)
        if val == None:
            val = -1
        result['lunch_break_finish_time'] = val

    row, col = find_marker_cell(sheet, 'Granularité')
    if row == None:
        logger.warning(f"The 'Granularité' cell in sheet {settings_sheet} is missing")
        duration = -1
    else:
        try:
            duration = int(parse_string(sheet, row, col + 1))
        except:
            duration = -1
    result['default_preference_duration'] = duration

    days = []
    row, col = find_marker_cell(sheet, 'Jours ouvrables')
    if row != None:
        # FIXME base.timing.Day has a CHOICES with this, but it's not available here
        for index, day in enumerate(['m', 'tu', 'w', 'th', 'f', 'sa', 'su']):
            if parse_string(sheet, row + 2, col + index) != '':
                days.append(day)
    result['days'] = days

    periods = dict()
    row, col = find_marker_cell(sheet, 'Périodes')
    if row != None:
        row = row + 2
        while row < REASONABLE:
            id_ = parse_string(sheet, row, col)
            if id_ == '':
                row = row + 1
                continue
            if id_ in periods:
                id_ = ':INVALID:DUPLICATE:{0:s}'.format(cell_name(row, col))
            start = parse_integer(sheet, row, col + 1)
            if start == None:
                start = -1
            finish = parse_integer(sheet, row, col + 2)
            if finish == None:
                finish = -1
            periods[id_] = (start, finish)
            row = row + 1
    result['periods'] = periods

    return result

def parse_groups(sheet):
    row_prom, col_prom = find_marker_cell(sheet, 'Identifiant')
    if row_prom == None:
        return dict(), set(), dict()

    row_nat, col_nat = find_marker_cell(sheet, 'Identifiant', row_prom + 1)
    if row_nat == None:
        return dict(), set(), dict()

    row_grp, col_grp = find_marker_cell(sheet, 'Identifiant', row_nat + 1)
    if row_grp == None:
        return dict(), set(), dict()

    promotions = dict()
    row = row_prom + 1
    while row < row_nat: # should stop before

        id_ = parse_string(sheet, row, col_prom)
        if id_ == '':
            break
        if id_ in promotions:
            id_ = ':INVALID:DUPLICATE:{0:s}'.format(cell_name(row, col_prom))
        promotions[id_] = parse_string(sheet, row, col_prom + 1)
        row = row + 1

    group_types = set()
    row = row_nat + 1
    while row < row_grp: # should stop before

        id_ = parse_string(sheet, row, col_nat)
        if id_ == '':
            break
        if id_ in group_types:
            id_ = ':INVALID:DUPLICATE:{0:s}'.format(cell_name(row, col_prom))
        group_types.add(id_)
        row = row + 1

    groups = dict()
    row = row_grp + 1
    while row < REASONABLE:

        id_ = parse_string(sheet, row, col_grp)
        if id_ == '':
            row = row + 1
            continue
        if id_ in groups:
            id_ = ':INVALID:DUPLICATE:{0:s}'.format(cell_name(row, col_prom))
        promotion = parse_string(sheet, row, col_grp + 1)
        group_type = parse_string(sheet, row, col_grp + 2)
        parent_ = parse_string(sheet, row, col_grp + 3)
        if parent_ == '':
            parent = set()
        else:
            parent = {parent_}
        groups[id_] = {'promotion': promotion,
                       'group_type' : group_type,
                       'parent' : parent}
        row = row + 1

    return promotions, group_types, groups


#################################################
#                                               #
#           Main parsing function               #
#                                               #
#################################################


def parse_file(filename = 'file_essai.xlsx'):
    try:
        wb = load_workbook(filename, data_only=True)

        sheet = wb[rooms_sheet]
        if not sheet:
            logger.warning(f"Sheet {rooms_sheet} doesn't exist")
            return None

        rooms, room_groups, room_categories = parse_rooms(sheet)

        sheet = wb[people_sheet]
        if not sheet:
            logger.warning(f"Sheet {people_sheet} doesn't exist")
            return None

        people = parse_people(sheet)

        sheet = wb[modules_sheet]
        if not sheet:
            logger.warning(f"Sheet {modules_sheet} doesn't exist")
            return None

        modules = parse_modules(sheet)

        sheet = wb[courses_sheet]
        if not sheet:
            logger.warning(f"Sheet {courses_sheet} doesn't exist")
            return None

        courses = parse_courses(sheet)

        sheet = wb[settings_sheet]
        if not sheet:
            logger.warning(f"Sheet {settings_sheet} doesn't exist")
            return None

        settings = parse_settings(sheet)

        sheet = wb[groups_sheet]
        if not sheet:
            logger.warning(f"Sheet {groups_sheet} doesn't exist")
            return None

        promotions, group_types, groups = parse_groups(sheet)

        return {'rooms' : rooms,
                'room_groups' : room_groups,
                'room_categories' : room_categories,
                'people' : people,
                'modules' : modules,
                'courses' : courses,
                'settings' : settings,
                'promotions': promotions,
                'group_types' : group_types,
                'groups' : groups }
    except FileNotFoundError as ex:
        logger.warning("Database file couldn't be opened: ", ex)
        return None

# dirty, but for testing purposes it's nice
if __name__ == '__main__':
    print("===== WARNINGS (should be empty) ======")
    results = parse_file()
    print("===== RESULTS ===========")
    print(results)
